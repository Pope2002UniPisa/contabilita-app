"""
Lettura/scrittura del template Excel di bilancio.
Calcola anche i totali (righe formula) direttamente in Python.
"""
import os
import shutil
import openpyxl
import config

_TEMPLATE_NAME = "Template_Bilancio.xlsx"
_WORKING_NAME  = "Bilancio.xlsx"

ANNO_COLONNE = {2022: "B", 2023: "C", 2024: "D", 2025: "E", 2026: "F", 2027: "G"}

# Dati grezzi del giornale → celle Excel
MAPPA = [
    ("CONTO ECONOMICO",  4,  ["80"],                      "netto_a"),
    ("CONTO ECONOMICO", 12,  ["60"],                      "netto_d"),
    ("CONTO ECONOMICO", 13,  ["61","62","63","64","65"],  "netto_d"),
    ("CONTO ECONOMICO", 14,  ["67","68"],                 "netto_d"),
    ("CONTO ECONOMICO", 28,  ["66","69"],                 "netto_d"),
    ("STATO PATRIMONIALE", 13, ["10"],                    "netto_d"),
    ("STATO PATRIMONIALE", 32, ["15"],                    "netto_d"),
    ("STATO PATRIMONIALE", 35, ["18"],                    "netto_d"),
    ("STATO PATRIMONIALE", 46, ["20"],                    "netto_d"),
    ("STATO PATRIMONIALE", 47, ["21"],                    "netto_d"),
    ("STATO PATRIMONIALE", 88, ["40"],                    "netto_a"),
    ("STATO PATRIMONIALE", 91, ["45","48"],               "netto_a"),
]

# Totali calcolati da Python (riga_tot, [righe_da_sommare])
# Queste sostituiscono le formule Excel che data_only=True non calcola
_SP_TOTALI = [
    (10, [7, 8, 9]),
    (17, [12, 13, 14, 15, 16]),
    (21, [20]),
    (22, [21]),
    (23, [10, 17, 22]),
    (29, [26, 27, 28]),
    (33, [32]),
    (36, [35]),
    (40, [39]),
    (41, [33, 36, 37, 40]),
    (44, [43]),
    (48, [46, 47]),
    (49, [29, 41, 44, 48]),
    (51, [4, 23, 49, 50]),
    (65, [62, 63, 64]),
    (71, [56, 57, 58, 59, 60, 65, 66, 67, 68, 69, 70]),
    (77, [73, 74, 75, 76]),
    (82, [81]),
    (86, [84, 85]),
    (89, [88]),
    (92, [91]),
    (95, [94]),
    (98, [97]),
    (99, [82, 86, 89, 92, 95, 98]),
    (101, [71, 77, 78, 99, 100]),
]

_CE_TOTALI = [
    (9,  [7, 8]),
    (10, [4, 5, 9]),
    (20, [16, 17, 18, 19]),
    (25, [22, 23, 24]),
    (29, [12, 13, 14, 20, 25, 26, 27, 28]),
    (35, [34]),
    (36, [35]),
    (39, [38]),
    (41, [36, 39, 40]),
    (45, [44]),
    (48, [47]),
    (49, [45, 48]),
    (54, [52, 53]),
]

FOGLI_DA_MOSTRARE = ["STATO PATRIMONIALE", "CONTO ECONOMICO", "INDICI",
                     "RENDICONTO FINANZIARIO", "VALUTAZIONE INVESTIMENTI", "WACC", "STAFF DISTRIBUTION"]


def _working_path():
    return os.path.join(config.get_data_dir(), _WORKING_NAME)


def _template_src():
    return os.path.join(config.get_resources_dir(), _TEMPLATE_NAME)


def assicura_excel():
    dst = _working_path()
    if not os.path.exists(dst):
        src = _template_src()
        if os.path.exists(src):
            os.makedirs(os.path.dirname(dst), exist_ok=True)
            shutil.copy2(src, dst)
    return os.path.exists(dst)


def excel_disponibile():
    return os.path.exists(_working_path())


def _calcola(saldi, conti, tipo):
    d = sum(saldi.get(c, {}).get("dare",  0.0) for c in conti)
    a = sum(saldi.get(c, {}).get("avere", 0.0) for c in conti)
    if tipo == "netto_d": return round(d - a, 2)
    if tipo == "netto_a": return round(a - d, 2)
    if tipo == "dare":    return round(d, 2)
    return round(a, 2)


def _calcola_totali(wb, col, foglio, totali_def):
    ws = wb[foglio]
    for riga_tot, righe_src in totali_def:
        val = 0.0
        for r in righe_src:
            cell_val = ws[f"{col}{r}"].value
            if isinstance(cell_val, (int, float)):
                val += cell_val
        if val != 0:
            ws[f"{col}{riga_tot}"] = round(val, 2)

    # Differenze speciali CE
    if foglio == "CONTO ECONOMICO":
        def _v(r):
            c = ws[f"{col}{r}"].value
            return c if isinstance(c, (int, float)) else 0.0
        # Riga 30: A - B
        diff_ab = round(_v(10) - _v(29), 2)
        ws[f"{col}30"] = diff_ab
        # Riga 50: (A-B) + C + D
        ws[f"{col}50"] = round(diff_ab + _v(41) + _v(49), 2)
        # Riga 55: Utile = 50 - imposte
        ws[f"{col}55"] = round(_v(50) - _v(54), 2)


def aggiorna_bilancio(saldi, anno):
    if anno not in ANNO_COLONNE:
        raise ValueError(f"Anno {anno} non supportato (2022-2027)")
    col = ANNO_COLONNE[anno]
    path = _working_path()
    wb = openpyxl.load_workbook(path)

    # 1) Scrivi i saldi grezzi
    cell_val: dict = {}
    for foglio, riga, conti, tipo in MAPPA:
        key = (foglio, riga)
        cell_val[key] = cell_val.get(key, 0.0) + _calcola(saldi, conti, tipo)
    for (foglio, riga), val in cell_val.items():
        if foglio in wb.sheetnames and val != 0:
            wb[foglio][f"{col}{riga}"] = round(val, 2)

    # 2) Calcola i totali (sostituisce formule Excel)
    if "STATO PATRIMONIALE" in wb.sheetnames:
        _calcola_totali(wb, col, "STATO PATRIMONIALE", _SP_TOTALI)
    if "CONTO ECONOMICO" in wb.sheetnames:
        _calcola_totali(wb, col, "CONTO ECONOMICO", _CE_TOTALI)

    wb.save(path)


def leggi_foglio(nome):
    path = _working_path()
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    if nome not in wb.sheetnames:
        return []
    ws = wb[nome]
    rows = []
    merged = set()
    for mr in ws.merged_cells.ranges:
        for r, c in mr.cells:
            if (r, c) != (mr.min_row, mr.min_col):
                merged.add((r, c))
    for row in ws.iter_rows():
        celle = []
        for cell in row:
            bg = bold = None
            val = None if (cell.row, cell.column) in merged else cell.value
            try:
                fill = cell.fill
                if fill.fill_type == "solid":
                    rgb = fill.fgColor.rgb
                    if rgb and rgb not in ("00000000","FF000000","00FFFFFF","FFFFFFFF"):
                        bg = "#" + rgb[-6:]
            except Exception:
                pass
            try:
                bold = bool(cell.font.bold)
            except Exception:
                bold = False
            indent = 0
            try:
                indent = int(cell.alignment.indent or 0)
            except Exception:
                pass
            celle.append({"value": val, "bg": bg, "bold": bold, "indent": indent})
        rows.append(celle)
    return rows
